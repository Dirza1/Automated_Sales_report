from openpyxl import load_workbook,Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference
from collections import defaultdict
import openpyxl
import logging


def main() -> None:
    """
    This script cleanes data from an incomming exel source and generates a KPI overview.
    """

    #First we need op open the file. We can do this in diferent ways. Lates file in a folder, open a window to manualy select. 
    #For this script we go with a hard coded name and location
    #We open this file in read only and data only for security and to ensure data is correctly transfered
    imput_excel:Workbook = load_workbook(filename="example_sales.xlsx",data_only=True)
    imput_data:Worksheet = imput_excel["Sheet1"]

    #We now setup logging. This allows us to use error logs to show if data needed to be cleaned up.
    logging.basicConfig(
    level=logging.INFO,                     
    format="%(asctime)s [%(levelname)s] %(message)s",  #
    datefmt="%Y-%m-%d %H:%M:%S",              
    filename="cleanup.log",                       
    filemode="w"                              
    )
    
    #we initiate two dictionaries to keep track of our KPI's

    customers:defaultdict = defaultdict(list)
    products:defaultdict = defaultdict(list)

    #Now that we have the data loaded we need to access it to "clean" it. Cleaning of data can be done any way we would like to do it.
    #Exclude data that is missing, add default values, calculate values or else. For this example we add default values were possible and ignore values where we cant.

    for col in imput_data.iter_cols(min_row=2):
        for cel in col:
            if cel.column != 1:
                break
            #we start with ignoring values we cant set to default
            if not isinstance(cel.value,int):
                logging.error(f"Missing OderID in row {cel.row}. This row was removed from the summary")
                continue
            if not isinstance(imput_data.cell(row=cel.row,column=5).value, int) or not isinstance(imput_data.cell(row=cel.row,column=6).value, (float,int)): # type: ignore
                logging.error(f"Missing either count or price in row {cel.row}. This row is removed from the summary")
                continue
            #This is where the cleanup happends
            date:str = imput_data.cell(row=cel.row,column=2).value if imput_data.cell(row=cel.row,column=2).value is not None else "1990-01-01" # type: ignore
            customer:str = imput_data.cell(row=cel.row,column=3).value if imput_data.cell(row=cel.row,column=3).value is not None else "Unknown" # type: ignore
            product:str = imput_data.cell(row=cel.row,column=4).value if imput_data.cell(row=cel.row,column=4).value is not None else "Unknown" # type: ignore
            count:int = imput_data.cell(row=cel.row,column=5).value # type: ignore
            price:float = imput_data.cell(row=cel.row,column=6).value # type: ignore
            total:float = count * price
            #logging is cleanup was required
            if date == "1990-01-01" or customer == "Unknown" or product == "Unknown":
                logging.warning(f"Missing data in row {cel.row}. Data was set to a default value")
            
            #we now have the clean data in memory for each row that the imput sheet has.
            #Now it is time to generate a report. We can generate a KPI report.
            #for this it is required that we set a few things in memory.
            #We will generate a customer KIP and an product KPI.

            

            if customer not in customers:
                customers[customer] = [1,total]
            else:
                customers[customer][0] += 1
                customers[customer][1] += total
            
            if product not in products:
                products[product] = [count,total]
            else:
                products[product][0] += count
                products[product][1] += total
            
    
    #we can now generate a new tab with this information in the same excel sheet.
    #we will also add a nice bar graph to the sheet voor data visualisation
    export_data_customer:Worksheet = imput_excel.create_sheet("Customer KPI's")
    count = 2

    export_data_customer.cell(row=1,column=1).value = "Customer"# type: ignore
    export_data_customer.cell(row=1,column=2).value = "Total Orders"# type: ignore
    export_data_customer.cell(row=1,column=3).value = "Total Value"# type: ignore

    for customer, value in customers.items():
        export_data_customer.cell(row=count,column=1).value = customer# type: ignore
        export_data_customer.cell(row=count,column=2).value = value[0]
        export_data_customer.cell(row=count,column=3).value = value[1]
        count += 1
    
    chart:BarChart = BarChart()
    chart.type = "col"
    chart.title = "Customer KPI's"
    chart.x_axis.title = "Customer" # type: ignore
    chart.y_axis.title = "Total" # type: ignore
    chart.legend = None

    data = Reference(export_data_customer,min_col=2, min_row=2, max_row=count-1,max_col=3)
    categories = Reference(export_data_customer, min_col=1, max_col=1,min_row=2,max_row=count-1)
    chart.add_data(data=data,titles_from_data=True)
    chart.set_categories(categories)

    export_data_customer.add_chart(chart, "E1")

    export_data_product:Worksheet = imput_excel.create_sheet("Product KPI's")
    count = 2

    export_data_product.cell(row=1,column=1).value = "Product"# type: ignore
    export_data_product.cell(row=1,column=2).value = "Total sold"# type: ignore
    export_data_product.cell(row=1,column=3).value = "Total Value" # type: ignore

    for product, value in products.items():
        export_data_product.cell(row=count,column=1).value = product # type: ignore
        export_data_product.cell(row=count,column=2).value = value[0]
        export_data_product.cell(row=count,column=3).value = value[1]
        count += 1
    
    chart:BarChart = BarChart()
    chart.type = "col"
    chart.title = "Product KPI's"
    chart.x_axis.title = "Product" # type: ignore
    chart.y_axis.title = "Total" # type: ignore
    chart.legend = None

    data = Reference(export_data_product,min_col=2, min_row=2, max_row=count-1,max_col=3)
    categories = Reference(export_data_product, min_col=1, max_col=1,min_row=2,max_row=count-1)
    chart.add_data(data=data,titles_from_data=True)
    chart.set_categories(categories)

    export_data_product.add_chart(chart, "E1")

    #then at the verry end we save the file

    imput_excel.save(filename="example_sales.xlsx")




if __name__ == "__main__":
    main()